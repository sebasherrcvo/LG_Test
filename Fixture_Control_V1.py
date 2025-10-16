"""Fixture Control System"""

import sys
import os
from datetime import datetime
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QLineEdit, QPushButton, 
                             QTableWidget, QTableWidgetItem, QMessageBox, 
                             QDialog, QFormLayout, QComboBox, QSpinBox, 
                             QTabWidget, QTextEdit, QHeaderView, QCheckBox)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QPalette, QColor
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
import glob
import shutil
import getpass


class FixtureControlApp(QMainWindow):
    def create_comparison_tab(self):
        """Create the fixture comparison tab"""
        comparison_widget = QWidget()
        layout = QVBoxLayout(comparison_widget)

        # Dropdown/search for fixture types
        type_select_layout = QHBoxLayout()
        type_label = QLabel("Select Fixture Type:")
        type_label.setFont(self.header_font)
        self.comparison_type_combo = QComboBox()
        self.comparison_type_combo.setEditable(True)
        self.comparison_type_combo.setFont(self.modern_font)
        types = getattr(self, 'types', []) if hasattr(self, 'types') else []
        self.comparison_type_combo.addItems(types)
        type_select_layout.addWidget(type_label)
        type_select_layout.addWidget(self.comparison_type_combo)
        layout.addLayout(type_select_layout)

        # Main content layout with image on left, details on right
        content_layout = QHBoxLayout()
        
        # Left side - Fixture Image
        image_container = QVBoxLayout()
        image_label_title = QLabel("Fixture Image:")
        image_label_title.setFont(self.header_font)
        image_container.addWidget(image_label_title)
        
        self.comparison_image_label = QLabel("No image available")
        self.comparison_image_label.setFixedSize(300, 300)
        self.comparison_image_label.setStyleSheet("border: 2px solid #0275d8; background-color: #f8f9fa; border-radius: 8px;")
        self.comparison_image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.comparison_image_label.setScaledContents(False)  # Don't stretch, we'll scale manually
        image_container.addWidget(self.comparison_image_label)
        image_container.addStretch()
        content_layout.addLayout(image_container)
        
        # Right side - Comparison data
        data_container = QVBoxLayout()
        self.qty_needed_label = QLabel("Quantity Needed: -")
        self.qty_needed_label.setFont(self.header_font)
        self.qty_actual_label = QLabel("Actual Quantity: -")
        self.qty_actual_label.setFont(self.header_font)
        self.qty_missing_label = QLabel("")
        self.qty_missing_label.setFont(self.header_font)
        self.qty_missing_label.setStyleSheet("color: red; font-weight: bold;")
        data_container.addWidget(self.qty_needed_label)
        data_container.addWidget(self.qty_actual_label)
        data_container.addWidget(self.qty_missing_label)
        data_container.addStretch()
        content_layout.addLayout(data_container)
        
        layout.addLayout(content_layout)


        def update_comparison():
            from PyQt6.QtGui import QPixmap
            import base64
            
            selected_type = self.comparison_type_combo.currentText()
            if not selected_type or selected_type.strip() == '':
                return  # Don't process empty selection
            
            selected_type_lower = selected_type.strip().lower()
            qty_needed = getattr(self, 'type_qty_needed', {}).get(selected_type, '-')
            
            # Update fixture image (with caching for better performance)
            type_images = getattr(self, 'type_images', {})
            if selected_type in type_images:
                try:
                    # Check cache first to avoid repeated base64 decoding
                    if selected_type in self._image_pixmap_cache:
                        scaled_pixmap = self._image_pixmap_cache[selected_type]
                        self.comparison_image_label.setPixmap(scaled_pixmap)
                    else:
                        # Decode and cache the image
                        image_base64 = type_images[selected_type]
                        image_bytes = base64.b64decode(image_base64)
                        
                        pixmap = QPixmap()
                        pixmap.loadFromData(image_bytes)
                        if not pixmap.isNull():
                            # Scale to fit while maintaining aspect ratio
                            scaled_pixmap = pixmap.scaled(
                                280, 280, 
                                Qt.AspectRatioMode.KeepAspectRatio,
                                Qt.TransformationMode.SmoothTransformation
                            )
                            # Cache the scaled pixmap for next time
                            self._image_pixmap_cache[selected_type] = scaled_pixmap
                            self.comparison_image_label.setPixmap(scaled_pixmap)
                        else:
                            self.comparison_image_label.clear()
                            self.comparison_image_label.setText("No image available")
                except Exception as e:
                    print(f"Error loading image: {e}")
                    self.comparison_image_label.clear()
                    self.comparison_image_label.setText("No image available")
            else:
                self.comparison_image_label.clear()
                self.comparison_image_label.setText("No image available")
            
            # Get actual quantity from cached inventory counts (avoid opening Excel repeatedly)
            actual_qty = getattr(self, '_comparison_inventory_cache', {}).get(selected_type_lower, 0)
            
            self.qty_needed_label.setText(f"Quantity Needed: {qty_needed}")
            self.qty_actual_label.setText(f"Actual Quantity: {actual_qty}")
            
            # Calculate and display missing fixtures only if there's a shortage
            if qty_needed != '-' and actual_qty != '-':
                missing = qty_needed - actual_qty
                if missing > 0:
                    self.qty_missing_label.setText(f"Missing: {missing}")
                    self.qty_missing_label.setVisible(True)
                else:
                    self.qty_missing_label.setText("")
                    self.qty_missing_label.setVisible(False)
            else:
                self.qty_missing_label.setText("")
                self.qty_missing_label.setVisible(False)

        # Build inventory cache function (called when tab opens)
        def build_inventory_cache():
            """Build a cache of fixture type counts to avoid repeated Excel reads"""
            try:
                from openpyxl import load_workbook
                self._comparison_inventory_cache = {}
                
                wb = load_workbook(self.excel_file, read_only=True, data_only=True)
                ws = wb.active
                
                # Count all fixture types in one pass
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[2]:  # Name column (fixture type)
                        fixture_type = str(row[2]).strip().lower()
                        self._comparison_inventory_cache[fixture_type] = self._comparison_inventory_cache.get(fixture_type, 0) + 1
                
                wb.close()
                print(f"Built comparison inventory cache with {len(self._comparison_inventory_cache)} fixture types")
            except Exception as e:
                print(f"Error building comparison inventory cache: {e}")
                self._comparison_inventory_cache = {}
        
        # Expose cache builder for external calls
        self.build_comparison_cache = build_inventory_cache
        
        self.comparison_type_combo.currentTextChanged.connect(update_comparison)
        # expose for external calls (for example when tab is opened)
        self.update_comparison_for_tab = update_comparison
        
        # Don't call update_comparison on init - wait until tab is opened
        # This prevents loading Excel file unnecessarily on startup

        # Store the comparison widget index for later use
        self.comparison_tab_widget = comparison_widget
        self.tabs.addTab(comparison_widget, "Fixture Comparison")
    def __init__(self):
        super().__init__()
        print("=" * 60)
        print("🚀 Starting Fixture Control System...")
        print("=" * 60)
        
        # Use shared network path for Excel file
        self.base_dir = r"\\slcnt005\Operations\Fixture Control\fixture_control"
        self.username = getpass.getuser()
        print(f"User: {self.username}")
        
        print("\n📁 Loading settings...")
        self.settings_file = self.get_most_recent_settings_file()
        self.models, self.types, self.type_to_models = self.load_settings()
        
        # Set Excel file path (file will be managed when saving data)
        self.excel_file = self.get_excel_file()
        print(f"\n📊 Excel file: {os.path.basename(self.excel_file)}")
        
        # On startup, identify and prepare the newest file for use (optimized)
        print("\n📋 Preparing Excel file...")
        self.prepare_excel_file_on_startup()
        
        # Image cache to avoid repeated base64 decoding (improves performance)
        self._image_pixmap_cache = {}
        
        # Comparison tab inventory cache to avoid repeated Excel reads
        self._comparison_inventory_cache = {}
        
        print("\n🎨 Initializing UI...")
        self.init_ui()
        
        # Load data if file exists (handled gracefully in load_data)
        print("\n📦 Loading inventory data...")
        self.load_data()
        
        print("\n✅ Fixture Control System ready!")
        print("=" * 60)

    def ensure_excel_file(self):
        """Ensure Excel file is ready - returns the path to {user}_inventory.xlsx"""
        return self.excel_file

    def get_most_recent_settings_file(self):
        """Get or create user's settings file using the most recent available (any user)"""
        import time
        # Find ALL .json files in the directory (not just specific naming patterns)
        settings_files = glob.glob(os.path.join(self.base_dir, "*.json"))
        user_file = os.path.join(self.base_dir, f"{self.username}_settings.json")
        
        if not settings_files:
            # No settings file exists, return user file (will be created later)
            print("No settings files found, will create new one")
            return user_file
        
        print(f"Found {len(settings_files)} JSON files: {[os.path.basename(f) for f in settings_files]}")
        
        # Filter out files we can't access (permission errors)
        accessible_files = []
        for f in settings_files:
            try:
                os.path.getmtime(f)
                accessible_files.append(f)
            except (PermissionError, OSError) as e:
                print(f"Skipping {os.path.basename(f)} - access denied")
        
        if not accessible_files:
            # No accessible files, return user file (will be created later)
            print("No accessible settings files found, will create new one")
            return user_file
        
        # Find most recent settings file (regardless of owner) and return it for reading
        most_recent = max(accessible_files, key=os.path.getmtime)
        most_recent_name = os.path.basename(most_recent)
        print(f"Most recent settings file: {most_recent_name} (will be used for loading)")
        # We intentionally DO NOT copy/rename here — load settings from the newest file available.
        return most_recent

    def get_newest_settings_file(self):
        """Dynamically find and return the newest settings JSON file path (regardless of owner or naming).
        Returns None if no JSON files found."""
        try:
            # Find ALL .json files in the directory
            settings_files = glob.glob(os.path.join(self.base_dir, "*.json"))
            
            if not settings_files:
                print(f"No JSON settings files found in {self.base_dir}")
                return None
            
            # Filter out files we can't access (permission errors)
            accessible_files = []
            for f in settings_files:
                try:
                    os.path.getmtime(f)
                    accessible_files.append(f)
                except (PermissionError, OSError):
                    pass  # Skip files we can't access
            
            if not accessible_files:
                print(f"No accessible JSON settings files found in {self.base_dir}")
                return None
            
            print(f"Found {len(accessible_files)} accessible JSON files: {[os.path.basename(f) for f in accessible_files]}")
            
            # Find newest by modification time
            newest_file = max(accessible_files, key=os.path.getmtime)
            print(f"Newest settings file: {os.path.basename(newest_file)}")
            return newest_file
        except Exception as e:
            print(f"Error finding newest settings file: {e}")
            return None

    def get_excel_file(self):
        """Get the path to the user_inventory.xlsx file"""
        # Simple approach: return the path to {user}_inventory.xlsx
        # File will be created by manage_excel_files_on_save if it doesn't exist
        return os.path.join(self.base_dir, f"{self.username}_inventory.xlsx")

    def get_newest_excel_file(self):
        """Dynamically find and return the newest Excel file path (regardless of owner or naming).
        Returns None if no Excel files found."""
        try:
            # Find ALL .xlsx files in the directory (not just specific naming patterns)
            excel_files = glob.glob(os.path.join(self.base_dir, "*.xlsx"))
            # Filter out temp files (starting with ~)
            excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~')]
            
            if not excel_files:
                print(f"No Excel files found in {self.base_dir}")
                return None
            
            # Filter out files we can't access (permission errors)
            accessible_files = []
            for f in excel_files:
                try:
                    os.path.getmtime(f)
                    accessible_files.append(f)
                except (PermissionError, OSError):
                    pass  # Skip files we can't access
            
            if not accessible_files:
                print(f"No accessible Excel files found in {self.base_dir}")
                return None
            
            print(f"Found {len(accessible_files)} accessible Excel files: {[os.path.basename(f) for f in accessible_files]}")
            
            # Find newest by modification time
            newest_file = max(accessible_files, key=os.path.getmtime)
            print(f"Newest Excel file: {os.path.basename(newest_file)}")
            return newest_file
        except Exception as e:
            print(f"Error finding newest Excel file: {e}")
            return None

    def load_settings(self):
        """Load settings from JSON file, create if doesn't exist"""
        import json
        import time
        
        if os.path.exists(self.settings_file):
            # Try to read settings file
            for attempt in range(5):
                try:
                    print(f"Loading settings from: {os.path.basename(self.settings_file)}...")
                    file_size = os.path.getsize(self.settings_file)
                    
                    # Warn if file is suspiciously large (likely has uncompressed images)
                    if file_size > 5_000_000:  # 5MB
                        print(f"⚠️  WARNING: Settings file is very large ({file_size/1024/1024:.1f} MB)")
                        print(f"⚠️  This may contain uncompressed images. Loading may be slow...")
                        print(f"⚠️  Tip: Re-upload images through Settings to compress them.")
                    
                    with open(self.settings_file, "r", encoding='utf-8') as f:
                        data = json.load(f)
                    
                    self.types = data.get("types", [])
                    self.models = data.get("models", [])
                    self.type_to_models = data.get("type_to_models", {})
                    self.type_descriptions = data.get("type_descriptions", {})
                    self.type_qty_needed = data.get("type_qty_needed", {})
                    
                    # Load WIP locations (default to standard locations if not present)
                    self.wip_locations = data.get("wip_locations", ["Bonepile", "Debug", "Operations", "Eng Lab"])
                    
                    # Load images but warn if there are many/large ones
                    self.type_images = data.get("type_images", {})
                    if self.type_images:
                        total_image_size = sum(len(img) for img in self.type_images.values())
                        print(f"Loaded {len(self.type_images)} image(s) ({total_image_size/1024:.0f} KB total)")
                        if total_image_size > 1_000_000:  # 1MB total
                            print(f"⚠️  Images are large. Consider re-uploading to compress them.")
                    
                    # Don't load images from Excel on startup - it's slow and unnecessary
                    # Images are already in JSON, Excel is just for viewing
                    # self.load_images_from_excel()  # REMOVED - causes slow startup
                    
                    print(f"✅ Settings loaded successfully")
                    return self.models, self.types, self.type_to_models
                    
                except PermissionError:
                    if attempt < 4:
                        print(f"Settings file locked, retrying... (attempt {attempt + 1}/5)")
                        time.sleep(1)
                    else:
                        print(f"❌ Settings file locked after 5 attempts, using defaults")
                        return [], [], {}
                        
                except json.JSONDecodeError as e:
                    print(f"❌ Settings file corrupted: {e}")
                    print(f"Creating backup and using defaults...")
                    try:
                        backup_file = self.settings_file + ".backup"
                        import shutil
                        shutil.copy2(self.settings_file, backup_file)
                        print(f"Backup saved to: {os.path.basename(backup_file)}")
                    except:
                        pass
                    break
                    
                except Exception as e:
                    print(f"❌ Error loading settings: {e}")
                    if attempt < 4:
                        time.sleep(1)
                    else:
                        return [], [], {}
        
        # File doesn't exist or is corrupted, create with defaults
        default_data = {"types": [], "models": [], "type_to_models": {}, "type_descriptions": {}}
        try:
            with open(self.settings_file, "w") as f:
                json.dump(default_data, f, indent=4)
        except Exception:
            pass
        
        return [], [], {}
        
    def init_excel_file(self):
        """Initialize Excel file if it doesn't exist - called on startup"""
        # Don't create the file on startup - let manage_excel_files_on_save handle it
        pass
    
    def manage_excel_files_on_save(self):
        """Consolidate all Excel files into {user}_inventory.xlsx when saving data.
        If current file is locked, create a timestamped working copy.
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            print(f"Managing Excel files for '{self.username}_inventory.xlsx' workflow...")
            current_path = self.excel_file  # This is {user}_inventory.xlsx
            processing_path = os.path.join(self.base_dir, f"{self.username}_inventory_processing.xlsx")
            
            # Find ALL Excel files in the directory (any naming pattern)
            excel_files = glob.glob(os.path.join(self.base_dir, "*.xlsx"))
            # Filter out temp files
            excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~')]
            print(f"Found {len(excel_files)} Excel files: {[os.path.basename(f) for f in excel_files]}")
            
            if not excel_files:
                # No Excel file found, create new one
                print(f"No Excel files found, creating new '{self.username}_inventory.xlsx'")
                wb = Workbook()
                ws = wb.active
                ws.title = "Inventory"
                headers = ["Serial", "Model", "Name", "Status", "Checked Out By", "Checked Out At", "Serialized Date", "WIP Location"]
                ws.append(headers)
                # Style headers
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                wb.save(current_path)
                print(f"Created new '{self.username}_inventory.xlsx'")
                self.excel_file = current_path
                return True
            else:
                # Excel file(s) found, preserve newest one's data
                excel_files_with_time = [(f, os.path.getmtime(f)) for f in excel_files]
                excel_files_with_time.sort(key=lambda x: x[1], reverse=True)  # Sort by time, newest first
                
                newest_file = excel_files_with_time[0][0]
                newest_name = os.path.basename(newest_file)
                
                print(f"Newest Excel file: {newest_name}")
                
                # Check if newest file is owned by current user
                is_owned = os.path.basename(newest_file).startswith(f"{self.username}_")
                
                if is_owned:
                    print(f"Newest file is owned by {self.username}")
                    
                    # If owned and not named current.xlsx, rename it
                    if newest_file != current_path:
                        try:
                            print(f"Renaming {os.path.basename(newest_file)} to {os.path.basename(current_path)}")
                            
                            # Delete current_path if it exists (it's older)
                            if os.path.exists(current_path):
                                try:
                                    os.remove(current_path)
                                    print(f"Deleted old {os.path.basename(current_path)}")
                                except Exception as e:
                                    print(f"Could not delete old current file: {e}")
                            
                            # Rename newest to current
                            os.rename(newest_file, current_path)
                            print(f"Renamed to {os.path.basename(current_path)}")
                            
                            # Update newest_file reference
                            newest_file = current_path
                            newest_name = os.path.basename(current_path)
                        except Exception as e:
                            print(f"Could not rename file: {e}")
                    
                    # Delete all other owned Excel files
                    for f in excel_files:
                        if f != newest_file and os.path.basename(f).startswith(f"{self.username}_"):
                            try:
                                os.remove(f)
                                print(f"Deleted other owned file: {os.path.basename(f)}")
                            except Exception as e:
                                print(f"Could not delete {os.path.basename(f)}: {e}")
                    
                    # Try to write to the file to see if it's locked
                    try:
                        wb_test = load_workbook(newest_file)
                        wb_test.save(newest_file)
                        wb_test.close()
                        print(f"{os.path.basename(newest_file)} is accessible and writable")
                        # Ensure final filename follows convention
                        desired = current_path
                        try:
                            if newest_file != desired:
                                # remove any existing desired file
                                if os.path.exists(desired):
                                    try:
                                        os.remove(desired)
                                    except Exception:
                                        pass
                                os.rename(newest_file, desired)
                                newest_file = desired
                                print(f"Renamed {os.path.basename(newest_file)} to {os.path.basename(desired)}")
                        except Exception as e:
                            print(f"Could not enforce final excel name: {e}")
                        self.excel_file = newest_file
                        return True
                    except Exception as e:
                        print(f"{os.path.basename(newest_file)} is locked: {e}")
                        # File is locked, create timestamped copy
                        import time
                        timestamp = time.strftime("%Y%m%d_%H%M%S")
                        timestamped_copy = os.path.join(self.base_dir, f"{self.username}_inventory_{timestamp}.xlsx")
                        
                        print(f"Creating timestamped working copy: {os.path.basename(timestamped_copy)}")
                        shutil.copy2(newest_file, timestamped_copy)
                        
                        # Verify writable
                        wb_verify = load_workbook(timestamped_copy)
                        wb_verify.save(timestamped_copy)
                        wb_verify.close()
                        print(f"Working copy is ready: {os.path.basename(timestamped_copy)}")
                        
                        self.excel_file = timestamped_copy
                        
                        # Clean up old timestamped copies
                        for f in excel_files:
                            if f != timestamped_copy and f != newest_file and os.path.basename(f).startswith(f"{self.username}_inventory_"):
                                try:
                                    os.remove(f)
                                    print(f"Deleted old timestamped copy: {os.path.basename(f)}")
                                except Exception:
                                    pass
                        
                        return True
                else:
                    print(f"Newest file is NOT owned by {self.username}, creating working copy")
                    
                    # Not owned, create timestamped copy from newest file
                    import time
                    timestamp = time.strftime("%Y%m%d_%H%M%S")
                    timestamped_copy = os.path.join(self.base_dir, f"{self.username}_current_{timestamp}.xlsx")
                    
                    print(f"Copying from {os.path.basename(newest_file)} to {os.path.basename(timestamped_copy)}")
                    shutil.copy2(newest_file, timestamped_copy)
                    
                    # Verify writable
                    wb_verify = load_workbook(timestamped_copy)
                    wb_verify.save(timestamped_copy)
                    wb_verify.close()
                    print(f"Working copy is ready: {os.path.basename(timestamped_copy)}")
                    
                    self.excel_file = timestamped_copy
                    
                    # Delete all other owned Excel files
                    for f in excel_files:
                        if f != timestamped_copy and os.path.basename(f).startswith(f"{self.username}_"):
                            try:
                                os.remove(f)
                                print(f"Deleted old owned file: {os.path.basename(f)}")
                            except Exception as e:
                                print(f"Could not delete {os.path.basename(f)}: {e}")
                    
                    return True
            
            print(f"Excel file management completed")
            return True
        except Exception as e:
            print(f"Error in manage_excel_files_on_save: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def prepare_excel_file_on_startup(self):
        """On startup, identify the newest Excel file and prepare it for use.
        This ensures security copies or other users' files are properly handled.
        """
        try:
            print(f"Preparing Excel file on startup...")
            current_path = self.excel_file  # This is {user}_inventory.xlsx
            
            # Find ALL Excel files in the directory (any naming pattern)
            excel_files = glob.glob(os.path.join(self.base_dir, "*.xlsx"))
            # Filter out temp files
            excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~')]
            
            if not excel_files:
                print(f"No Excel files found on startup")
                return
            
            print(f"Found {len(excel_files)} Excel files: {[os.path.basename(f) for f in excel_files]}")
            
            # Find the newest file (with permission error handling)
            excel_files_with_time = []
            for f in excel_files:
                try:
                    mtime = os.path.getmtime(f)
                    excel_files_with_time.append((f, mtime))
                except (PermissionError, OSError) as e:
                    print(f"Skipping {os.path.basename(f)} - access denied")
            
            if not excel_files_with_time:
                print("No accessible Excel files found")
                return
                
            excel_files_with_time.sort(key=lambda x: x[1], reverse=True)
            
            newest_file = excel_files_with_time[0][0]
            newest_name = os.path.basename(newest_file)
            
            print(f"Newest Excel file: {newest_name}")
            
            # We will always load data from the newest Excel file available (regardless of owner).
            # To allow writing, create/overwrite a per-user copy at current_path using newest as source.
            try:
                # Check if newest file is already the per-user file
                if os.path.abspath(newest_file) == os.path.abspath(current_path):
                    print(f"Newest Excel is already the per-user inventory: {os.path.basename(current_path)}")
                    self.excel_file = current_path
                else:
                    print(f"Copying newest Excel ({newest_name}) to per-user inventory: {os.path.basename(current_path)}")
                    # Remove existing per-user file if present
                    if os.path.exists(current_path):
                        try:
                            os.remove(current_path)
                        except Exception:
                            pass
                    shutil.copy2(newest_file, current_path)
                    # Skip verification on startup - it's slow and not critical
                    # File will be verified when user actually saves data
                    self.excel_file = current_path
                    print(f"✅ Per-user inventory ready: {os.path.basename(current_path)}")

                # After creating per-user copy, delete other owned files for this user (safe cleanup)
                for f in excel_files:
                    if f != self.excel_file and os.path.basename(f).startswith(f"{self.username}_"):
                        try:
                            os.remove(f)
                            print(f"Deleted other owned file: {os.path.basename(f)}")
                        except Exception as e:
                            print(f"Could not delete {os.path.basename(f)}: {e}")

            except Exception as e:
                print(f"Could not prepare per-user excel from newest file: {e}")
                # fallback: use newest file path directly
                self.excel_file = newest_file
            
            print(f"Startup: Using Excel file: {os.path.basename(self.excel_file)}")
            
        except Exception as e:
            print(f"Error in prepare_excel_file_on_startup: {e}")
            import traceback
            traceback.print_exc()
    

    def init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle("Fixture Control System")
        self.setGeometry(100, 100, 1200, 700)

        # Modern font
        from PyQt6.QtGui import QFont
        self.modern_font = QFont("Segoe UI", 14, QFont.Weight.Bold)
        self.header_font = QFont("Segoe UI", 16, QFont.Weight.Black)

        # Menu bar
        menubar = self.menuBar()
        settings_menu = menubar.addMenu("Settings")
        edit_settings_action = settings_menu.addAction("Edit Models/Names")
        edit_settings_action.triggered.connect(self.open_settings_dialog)
        barcode_menu = menubar.addMenu("Barcode")
        barcode_action = barcode_menu.addAction("Generate & Print Barcode")
        barcode_action.triggered.connect(self.open_barcode_menu)

        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Create tab widget
        self.tabs = QTabWidget()
        self.tabs.setFont(self.header_font)
        main_layout.addWidget(self.tabs)

        # Create tabs
        self.create_inventory_tab()
        self.create_checkin_checkout_tab()
        self.create_serialize_tab()
        self.create_comparison_tab()
        
        # Centralized tab change handler (prevents duplicate connections)
        self.tabs.currentChanged.connect(self._on_tab_changed)

        # Timer to refresh inventory data and settings (optimized interval)
        self.refresh_timer = QTimer(self)
        self.refresh_timer.timeout.connect(self.auto_refresh)
        self.refresh_timer.start(30000)  # 30 seconds (reduced from 5s for performance)
        
        # Track file modification times to avoid unnecessary reloads
        self._last_excel_mtime = 0
        self._last_settings_mtime = 0
        self._cached_excel_data = None
    
    def _on_tab_changed(self, index):
        """Centralized handler for tab changes"""
        # Refresh comparison tab when it's opened
        if hasattr(self, 'comparison_tab_widget') and self.tabs.widget(index) == self.comparison_tab_widget:
            try:
                # Reload settings from newest file to pick up qty changes
                self.models, self.types, self.type_to_models = self.reload_settings_file()
                
                # Build inventory cache ONCE when tab opens (instead of on every dropdown change)
                if hasattr(self, 'build_comparison_cache'):
                    self.build_comparison_cache()
                
                # Refresh the dropdown with updated types
                if hasattr(self, 'comparison_type_combo'):
                    current_selection = self.comparison_type_combo.currentText()
                    self.comparison_type_combo.clear()
                    self.comparison_type_combo.addItems(self.types)
                    # Try to restore previous selection if it still exists
                    if current_selection in self.types:
                        self.comparison_type_combo.setCurrentText(current_selection)
                
                # Update comparison display
                if hasattr(self, 'update_comparison_for_tab'):
                    self.update_comparison_for_tab()
            except Exception as e:
                print(f"Error refreshing comparison tab: {e}")
        
        # Auto-focus serial input when Serialize tab is opened
        elif hasattr(self, 'serialize_tab_widget') and self.tabs.widget(index) == self.serialize_tab_widget:
            if hasattr(self, 'new_serial_input'):
                self.new_serial_input.setFocus()
                self.new_serial_input.selectAll()  # Select any existing text for easy replacement

    def open_settings_dialog(self):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QListWidget, QLineEdit, QPushButton, QHBoxLayout, QComboBox
        import json
        from PyQt6.QtWidgets import QTabWidget
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Models and Fixture Types")
        dialog.resize(700, 600)
        tabs = QTabWidget(dialog)
        main_layout = QVBoxLayout(dialog)
        main_layout.addWidget(tabs)

        # Track name changes for Excel updates
        model_name_changes = {}  # old_name: new_name
        type_name_changes = {}   # old_name: new_name
        
        # --- Models Tab ---
        models_tab = QWidget()
        models_layout = QVBoxLayout(models_tab)
        models_layout.addWidget(QLabel("Models:"))
        
        # Add search box for models
        models_search = QLineEdit()
        models_search.setPlaceholderText("Search models...")
        models_layout.addWidget(models_search)
        
        self.models_list = QListWidget()
        models = getattr(self, 'models', []) if hasattr(self, 'models') else []
        self.models_list.addItems(models)
        models_layout.addWidget(self.models_list)
        
        # Filter function for models
        def filter_models():
            search_text = models_search.text().lower()
            for i in range(self.models_list.count()):
                item = self.models_list.item(i)
                item.setHidden(search_text not in item.text().lower())
        models_search.textChanged.connect(filter_models)
        model_input = QLineEdit()
        model_input.setPlaceholderText("Add new model...")
        add_model_btn = QPushButton("Add Model")
        def add_model():
            text = model_input.text().strip()
            if text and text not in models:
                self.models_list.addItem(text)
                models.append(text)
                model_input.clear()
        add_model_btn.clicked.connect(add_model)
        
        edit_model_btn = QPushButton("Edit Selected Model")
        def edit_model():
            selected = self.models_list.selectedItems()
            if selected:
                old_name = selected[0].text()
                from PyQt6.QtWidgets import QInputDialog
                new_name, ok = QInputDialog.getText(dialog, "Edit Model", f"Rename '{old_name}' to:", text=old_name)
                if ok and new_name.strip() and new_name.strip() != old_name:
                    new_name = new_name.strip()
                    # Update list
                    selected[0].setText(new_name)
                    # Update models array
                    idx = models.index(old_name)
                    models[idx] = new_name
                    # Track change for Excel update
                    model_name_changes[old_name] = new_name
                    # Update type_to_models mappings
                    for type_name, model_list in type_to_models.items():
                        if old_name in model_list:
                            type_to_models[type_name] = [new_name if m == old_name else m for m in model_list]
        edit_model_btn.clicked.connect(edit_model)
        
        remove_model_btn = QPushButton("Remove Selected Model")
        def remove_model():
            for item in self.models_list.selectedItems():
                m = item.text()
                self.models_list.takeItem(self.models_list.row(item))
                if m in models:
                    models.remove(m)
        remove_model_btn.clicked.connect(remove_model)
        model_btn_layout = QHBoxLayout()
        model_btn_layout.addWidget(model_input)
        model_btn_layout.addWidget(add_model_btn)
        model_btn_layout.addWidget(edit_model_btn)
        model_btn_layout.addWidget(remove_model_btn)
        models_layout.addLayout(model_btn_layout)
        tabs.addTab(models_tab, "Models")

        # --- Types Tab ---
        types_tab = QWidget()
        types_layout = QVBoxLayout(types_tab)
        types_layout.addWidget(QLabel("Fixture Types:"))
        
        # Add search box for fixture types
        types_search = QLineEdit()
        types_search.setPlaceholderText("Search fixture types...")
        types_layout.addWidget(types_search)
        
        self.types_list = QListWidget()
        types = getattr(self, 'types', []) if hasattr(self, 'types') else []
        self.types_list.addItems(types)
        types_layout.addWidget(self.types_list)
        
        # Filter function for types
        def filter_types():
            search_text = types_search.text().lower()
            for i in range(self.types_list.count()):
                item = self.types_list.item(i)
                item.setHidden(search_text not in item.text().lower())
        types_search.textChanged.connect(filter_types)

        types_layout.addWidget(QLabel("Type Description (editable):"))
        type_desc_input = QLineEdit()
        type_desc_input.setPlaceholderText("Enter a description for the selected type")
        types_layout.addWidget(type_desc_input)

        # Quantity needed input for types
        types_layout.addWidget(QLabel("Quantity Needed:"))
        qty_needed_input = QSpinBox()
        qty_needed_input.setMinimum(0)
        qty_needed_input.setMaximum(10000)
        qty_needed_input.setValue(0)
        types_layout.addWidget(qty_needed_input)
        
        # Image upload section
        types_layout.addWidget(QLabel("Fixture Type Image:"))
        image_layout = QHBoxLayout()
        image_preview_label = QLabel("No image")
        image_preview_label.setFixedSize(150, 150)
        image_preview_label.setStyleSheet("border: 1px solid gray; background-color: #f0f0f0;")
        image_preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        image_preview_label.setScaledContents(True)
        image_layout.addWidget(image_preview_label)
        
        image_buttons_layout = QVBoxLayout()
        upload_image_btn = QPushButton("Upload Image")
        remove_image_btn = QPushButton("Remove Image")
        image_buttons_layout.addWidget(upload_image_btn)
        image_buttons_layout.addWidget(remove_image_btn)
        image_buttons_layout.addStretch()
        image_layout.addLayout(image_buttons_layout)
        types_layout.addLayout(image_layout)
        
        # Store for current image data
        current_image_data = {'data': None}
        
        def upload_image():
            from PyQt6.QtWidgets import QFileDialog
            from PyQt6.QtGui import QPixmap
            import base64
            
            selected = self.types_list.selectedItems()
            if not selected:
                QMessageBox.warning(dialog, "No Selection", "Please select a fixture type first.")
                return
            
            file_path, _ = QFileDialog.getOpenFileName(
                dialog, "Select Image", "", "Image Files (*.png *.jpg *.jpeg *.bmp *.gif)"
            )
            if file_path:
                try:
                    from PIL import Image
                    from io import BytesIO
                    
                    # Open and resize image to max 400x400 to reduce file size
                    img = Image.open(file_path)
                    
                    # Convert RGBA to RGB if needed (for JPEG compatibility)
                    if img.mode in ('RGBA', 'LA', 'P'):
                        background = Image.new('RGB', img.size, (255, 255, 255))
                        if img.mode == 'P':
                            img = img.convert('RGBA')
                        background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                        img = background
                    
                    # Resize to max 400x400 (maintains aspect ratio)
                    img.thumbnail((400, 400), Image.Resampling.LANCZOS)
                    
                    # Save as compressed JPEG to reduce size
                    buffer = BytesIO()
                    img.save(buffer, format='JPEG', quality=85, optimize=True)
                    image_bytes = buffer.getvalue()
                    
                    # Convert to base64 for storage
                    image_base64 = base64.b64encode(image_bytes).decode('utf-8')
                    current_image_data['data'] = image_base64
                    
                    print(f"Image compressed: Original vs Compressed = {len(open(file_path, 'rb').read())} vs {len(image_bytes)} bytes")
                    
                    # Display preview
                    pixmap = QPixmap(file_path)
                    if not pixmap.isNull():
                        image_preview_label.setPixmap(pixmap.scaled(
                            150, 150, Qt.AspectRatioMode.KeepAspectRatio, 
                            Qt.TransformationMode.SmoothTransformation
                        ))
                    
                    # Store in type_images dict
                    if not hasattr(self, 'type_images'):
                        self.type_images = {}
                    self.type_images[selected[0].text()] = image_base64
                    
                except Exception as e:
                    QMessageBox.warning(dialog, "Error", f"Failed to load image: {str(e)}")
        
        def remove_image():
            selected = self.types_list.selectedItems()
            if selected:
                t = selected[0].text()
                if hasattr(self, 'type_images') and t in self.type_images:
                    del self.type_images[t]
                current_image_data['data'] = None
                image_preview_label.clear()
                image_preview_label.setText("No image")
        
        upload_image_btn.clicked.connect(upload_image)
        remove_image_btn.clicked.connect(remove_image)

        # When a type is selected, populate the description, qty, and image fields
        def on_type_selected():
            from PyQt6.QtGui import QPixmap
            import base64
            from io import BytesIO
            
            selected = self.types_list.selectedItems()
            if selected:
                t = selected[0].text()
                desc = getattr(self, 'type_descriptions', {}).get(t, '')
                type_desc_input.setText(desc)
                qty_needed = getattr(self, 'type_qty_needed', {}).get(t, 0)
                qty_needed_input.setValue(qty_needed)
                
                # Load image if exists
                if hasattr(self, 'type_images') and t in self.type_images:
                    try:
                        image_base64 = self.type_images[t]
                        current_image_data['data'] = image_base64
                        image_bytes = base64.b64decode(image_base64)
                        
                        pixmap = QPixmap()
                        pixmap.loadFromData(image_bytes)
                        if not pixmap.isNull():
                            image_preview_label.setPixmap(pixmap.scaled(
                                150, 150, Qt.AspectRatioMode.KeepAspectRatio,
                                Qt.TransformationMode.SmoothTransformation
                            ))
                        else:
                            image_preview_label.clear()
                            image_preview_label.setText("No image")
                    except Exception as e:
                        print(f"Error loading image: {e}")
                        image_preview_label.clear()
                        image_preview_label.setText("No image")
                        current_image_data['data'] = None
                else:
                    image_preview_label.clear()
                    image_preview_label.setText("No image")
                    current_image_data['data'] = None
            else:
                type_desc_input.clear()
                qty_needed_input.setValue(0)
                image_preview_label.clear()
                image_preview_label.setText("No image")
                current_image_data['data'] = None
        self.types_list.itemSelectionChanged.connect(on_type_selected)

        type_input = QLineEdit()
        type_input.setPlaceholderText("Add new fixture type...")
        self.type_model_binding_list = QListWidget()
        self.type_model_binding_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        type_to_models = getattr(self, 'type_to_models', {}) if hasattr(self, 'type_to_models') else {}
        def add_type():
            text = type_input.text().strip()
            if text and text not in types:
                types.append(text)
                self.types_list.addItem(text)
                # capture description if provided
                desc = type_desc_input.text().strip()
                if not hasattr(self, 'type_descriptions'):
                    self.type_descriptions = {}
                if desc:
                    self.type_descriptions[text] = desc
                # capture qty needed
                qty_needed = qty_needed_input.value()
                if not hasattr(self, 'type_qty_needed'):
                    self.type_qty_needed = {}
                self.type_qty_needed[text] = qty_needed
                type_input.clear()
                type_desc_input.clear()
                qty_needed_input.setValue(0)
        add_type_btn = QPushButton("Add Type")
        add_type_btn.clicked.connect(add_type)
        
        edit_type_btn = QPushButton("Edit Selected Type")
        def edit_type():
            selected = self.types_list.selectedItems()
            if selected:
                old_name = selected[0].text()
                from PyQt6.QtWidgets import QInputDialog
                new_name, ok = QInputDialog.getText(dialog, "Edit Type", f"Rename '{old_name}' to:", text=old_name)
                if ok and new_name.strip() and new_name.strip() != old_name:
                    new_name = new_name.strip()
                    # Update list
                    selected[0].setText(new_name)
                    # Update types array
                    idx = types.index(old_name)
                    types[idx] = new_name
                    # Track change for Excel update
                    type_name_changes[old_name] = new_name
                    # Update type_to_models mappings
                    if old_name in type_to_models:
                        type_to_models[new_name] = type_to_models.pop(old_name)
                    # Update combo box
                    self.type_model_combo.clear()
                    self.type_model_combo.addItems(types)
                    # Move description if present
                    if hasattr(self, 'type_descriptions') and old_name in self.type_descriptions:
                        self.type_descriptions[new_name] = self.type_descriptions.pop(old_name)
                    # Move qty_needed if present
                    if hasattr(self, 'type_qty_needed') and old_name in self.type_qty_needed:
                        self.type_qty_needed[new_name] = self.type_qty_needed.pop(old_name)
                    # Move image if present
                    if hasattr(self, 'type_images') and old_name in self.type_images:
                        self.type_images[new_name] = self.type_images.pop(old_name)
        edit_type_btn.clicked.connect(edit_type)
        
        remove_type_btn = QPushButton("Remove Selected Type")
        def remove_type():
            for item in self.types_list.selectedItems():
                t = item.text()
                self.types_list.takeItem(self.types_list.row(item))
                if t in types:
                    types.remove(t)
                    if t in type_to_models:
                        del type_to_models[t]
                    # remove description too
                    if hasattr(self, 'type_descriptions') and t in self.type_descriptions:
                        del self.type_descriptions[t]
                    # remove qty_needed too
                    if hasattr(self, 'type_qty_needed') and t in self.type_qty_needed:
                        del self.type_qty_needed[t]
                    # remove image too
                    if hasattr(self, 'type_images') and t in self.type_images:
                        del self.type_images[t]
        remove_type_btn.clicked.connect(remove_type)
        type_btn_layout = QHBoxLayout()
        type_btn_layout.addWidget(type_input)
        type_btn_layout.addWidget(add_type_btn)
        type_btn_layout.addWidget(edit_type_btn)
        type_btn_layout.addWidget(remove_type_btn)
        types_layout.addLayout(type_btn_layout)

        # Type-to-model binding section (dual list for batch add/remove)
        types_layout.addWidget(QLabel("Edit Type-to-Model Bindings:"))
        self.type_model_combo = QComboBox()
        self.type_model_combo.setEditable(True)  # Make it searchable
        self.type_model_combo.addItems(types)
        types_layout.addWidget(self.type_model_combo)
        
        binding_layout = QHBoxLayout()
        
        # Left side - Bound Models with search
        bound_container = QVBoxLayout()
        bound_container.addWidget(QLabel("Bound Models"))
        bound_search = QLineEdit()
        bound_search.setPlaceholderText("Search bound models...")
        bound_container.addWidget(bound_search)
        self.bound_models_list = QListWidget()
        self.bound_models_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        bound_container.addWidget(self.bound_models_list)
        binding_layout.addLayout(bound_container)
        
        # Filter function for bound models
        def filter_bound_models():
            search_text = bound_search.text().lower()
            for i in range(self.bound_models_list.count()):
                item = self.bound_models_list.item(i)
                item.setHidden(search_text not in item.text().lower())
        bound_search.textChanged.connect(filter_bound_models)
        
        # Middle buttons
        btn_layout = QVBoxLayout()
        add_btn = QPushButton("Add →")
        remove_btn = QPushButton("← Remove")
        btn_layout.addStretch()
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(remove_btn)
        btn_layout.addStretch()
        binding_layout.addLayout(btn_layout)
        
        # Right side - Unbound Models with search
        unbound_container = QVBoxLayout()
        unbound_container.addWidget(QLabel("Unbound Models"))
        unbound_search = QLineEdit()
        unbound_search.setPlaceholderText("Search unbound models...")
        unbound_container.addWidget(unbound_search)
        self.unbound_models_list = QListWidget()
        self.unbound_models_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        unbound_container.addWidget(self.unbound_models_list)
        binding_layout.addLayout(unbound_container)
        
        # Filter function for unbound models
        def filter_unbound_models():
            search_text = unbound_search.text().lower()
            for i in range(self.unbound_models_list.count()):
                item = self.unbound_models_list.item(i)
                item.setHidden(search_text not in item.text().lower())
        unbound_search.textChanged.connect(filter_unbound_models)
        
        types_layout.addLayout(binding_layout)

        def update_binding_lists():
            # Clear search boxes when switching types
            bound_search.clear()
            unbound_search.clear()
            self.bound_models_list.clear()
            self.unbound_models_list.clear()
            selected_type = self.type_model_combo.currentText()
            bound = set(type_to_models.get(selected_type, []))
            for m in models:
                if m in bound:
                    self.bound_models_list.addItem(m)
                else:
                    self.unbound_models_list.addItem(m)
        self.type_model_combo.currentTextChanged.connect(update_binding_lists)
        update_binding_lists()

        def add_models():
            selected_type = self.type_model_combo.currentText()
            selected = [item.text() for item in self.unbound_models_list.selectedItems()]
            current = set(type_to_models.get(selected_type, []))
            current.update(selected)
            type_to_models[selected_type] = list(current)
            update_binding_lists()
        add_btn.clicked.connect(add_models)

        def remove_models():
            selected_type = self.type_model_combo.currentText()
            selected = [item.text() for item in self.bound_models_list.selectedItems()]
            current = set(type_to_models.get(selected_type, []))
            current.difference_update(selected)
            type_to_models[selected_type] = list(current)
            update_binding_lists()
        remove_btn.clicked.connect(remove_models)

        tabs.addTab(types_tab, "Fixture Types")

        # --- WIP Locations Tab ---
        wip_locations_tab = QWidget()
        wip_locations_layout = QVBoxLayout(wip_locations_tab)
        wip_locations_layout.addWidget(QLabel("WIP Locations:"))
        wip_locations_layout.addWidget(QLabel("(Locations used for Work In Progress tagging)"))
        
        # Add search box for WIP locations
        wip_search = QLineEdit()
        wip_search.setPlaceholderText("Search WIP locations...")
        wip_locations_layout.addWidget(wip_search)
        
        self.wip_locations_list = QListWidget()
        wip_locations = getattr(self, 'wip_locations', ["Bonepile", "Debug", "Operations", "Eng Lab"])
        self.wip_locations_list.addItems(wip_locations)
        wip_locations_layout.addWidget(self.wip_locations_list)
        
        # Filter function for WIP locations
        def filter_wip_locations():
            search_text = wip_search.text().lower()
            for i in range(self.wip_locations_list.count()):
                item = self.wip_locations_list.item(i)
                item.setHidden(search_text not in item.text().lower())
        wip_search.textChanged.connect(filter_wip_locations)
        
        wip_location_input = QLineEdit()
        wip_location_input.setPlaceholderText("Add new WIP location...")
        add_wip_location_btn = QPushButton("Add Location")
        def add_wip_location():
            text = wip_location_input.text().strip()
            if text and text not in wip_locations:
                self.wip_locations_list.addItem(text)
                wip_locations.append(text)
                wip_location_input.clear()
                # Update the wip_location_combo in Check In/Out tab
                if hasattr(self, 'wip_location_combo'):
                    self.wip_location_combo.clear()
                    self.wip_location_combo.addItems(wip_locations)
        add_wip_location_btn.clicked.connect(add_wip_location)
        
        edit_wip_location_btn = QPushButton("Edit Selected Location")
        def edit_wip_location():
            selected = self.wip_locations_list.selectedItems()
            if selected:
                old_name = selected[0].text()
                from PyQt6.QtWidgets import QInputDialog
                new_name, ok = QInputDialog.getText(dialog, "Edit WIP Location", f"Rename '{old_name}' to:", text=old_name)
                if ok and new_name.strip() and new_name.strip() != old_name:
                    new_name = new_name.strip()
                    # Update list
                    selected[0].setText(new_name)
                    # Update wip_locations array
                    idx = wip_locations.index(old_name)
                    wip_locations[idx] = new_name
                    # Update combo box in Check In/Out tab
                    if hasattr(self, 'wip_location_combo'):
                        self.wip_location_combo.clear()
                        self.wip_location_combo.addItems(wip_locations)
        edit_wip_location_btn.clicked.connect(edit_wip_location)
        
        remove_wip_location_btn = QPushButton("Remove Selected Location")
        def remove_wip_location():
            for item in self.wip_locations_list.selectedItems():
                loc = item.text()
                self.wip_locations_list.takeItem(self.wip_locations_list.row(item))
                if loc in wip_locations:
                    wip_locations.remove(loc)
                    # Update combo box in Check In/Out tab
                    if hasattr(self, 'wip_location_combo'):
                        self.wip_location_combo.clear()
                        self.wip_location_combo.addItems(wip_locations)
        remove_wip_location_btn.clicked.connect(remove_wip_location)
        
        wip_btn_layout = QHBoxLayout()
        wip_btn_layout.addWidget(wip_location_input)
        wip_btn_layout.addWidget(add_wip_location_btn)
        wip_btn_layout.addWidget(edit_wip_location_btn)
        wip_btn_layout.addWidget(remove_wip_location_btn)
        wip_locations_layout.addLayout(wip_btn_layout)
        tabs.addTab(wip_locations_tab, "WIP Locations")

        # Save button (single)
        save_btn = QPushButton("Save Settings")
        def save_settings():
            import json
            # Ensure type_descriptions and type_qty_needed exist and are updated from UI
            type_descriptions = getattr(self, 'type_descriptions', {})
            type_qty_needed = getattr(self, 'type_qty_needed', {})
            # If a type is selected and description/qty edited, persist it
            selected = self.types_list.selectedItems()
            if selected:
                sel = selected[0].text()
                d = type_desc_input.text().strip()
                if d:
                    type_descriptions[sel] = d
                qty_needed = qty_needed_input.value()
                type_qty_needed[sel] = qty_needed

            # Get type_images dict (contains base64 encoded images)
            type_images = getattr(self, 'type_images', {})
            
            # Always save settings to the user's own settings file to avoid overwriting others' newest file
            user_settings_path = os.path.join(self.base_dir, f"{self.username}_settings.json")
            with open(user_settings_path, "w", encoding='utf-8') as f:
                json.dump({
                    "types": types, 
                    "models": models, 
                    "type_to_models": type_to_models, 
                    "type_descriptions": type_descriptions, 
                    "type_qty_needed": type_qty_needed,
                    "type_images": type_images,
                    "wip_locations": wip_locations
                }, f, indent=2)
            # Update self.settings_file to point to the user's file
            self.settings_file = user_settings_path
            # store on self
            self.type_descriptions = type_descriptions
            self.type_qty_needed = type_qty_needed
            self.type_images = type_images
            self.types = types
            self.models = models
            self.type_to_models = type_to_models
            self.wip_locations = wip_locations
            
            # Clear image cache since images may have changed
            self._image_pixmap_cache = {}
            
            # Save images to Excel file
            self.save_images_to_excel(type_images)
            
            # Apply name changes to Excel file
            if model_name_changes or type_name_changes:
                self.update_excel_names(model_name_changes, type_name_changes)
            
            # Re-read settings from THE USER'S OWN FILE (not newest) to prevent data loss
            # DO NOT call reload_settings_file() as it loads from newest file
            # The data is already stored in memory (self.types, self.models, etc.)
            # Refresh UI elements with the saved data
            self.models_list.clear()
            self.models_list.addItems(self.models)
            self.types_list.clear()
            self.types_list.addItems(self.types)
            self.type_model_combo.clear()
            self.type_model_combo.addItems(self.types)
            update_binding_lists()
            # Also update serializer tab drop-downs
            if hasattr(self, 'update_serializer_dropdowns'):
                self.update_serializer_dropdowns()
            # Update comparison tab dropdown with new types
            if hasattr(self, 'comparison_type_combo'):
                current_selection = self.comparison_type_combo.currentText()
                self.comparison_type_combo.clear()
                self.comparison_type_combo.addItems(self.types)
                # Restore selection if it still exists
                if current_selection in self.types:
                    self.comparison_type_combo.setCurrentText(current_selection)
            # Update Excel logs to match compatible models for each fixture type
            self.update_excel_model_compatibility()
            # Refresh inventory display
            self.load_data()
            # Trigger comparison update if the function exists
            if hasattr(self, 'update_comparison_for_tab'):
                self.update_comparison_for_tab()
            QMessageBox.information(None, "Settings Saved", "Settings were successfully saved and inventory updated.")
        save_btn.clicked.connect(save_settings)
        main_layout.addWidget(save_btn)

        dialog.exec()

    def reload_settings_file(self):
        import json
        # Always reload from the newest settings file available
        newest_settings = self.get_newest_settings_file()
        if newest_settings and os.path.exists(newest_settings):
            print(f"Reloading settings from newest file: {os.path.basename(newest_settings)}")
            try:
                with open(newest_settings, "r") as f:
                    data = json.load(f)
                types = data.get("types", [])
                models = data.get("models", [])
                type_to_models = data.get("type_to_models", {})
                # load type descriptions if present
                self.type_descriptions = data.get("type_descriptions", {})
                # Load quantity needed data
                self.type_qty_needed = data.get("type_qty_needed", {})
                # Load images data
                self.type_images = data.get("type_images", {})
                # Also load images from Excel
                self.load_images_from_excel()
                return models, types, type_to_models
            except Exception as e:
                print(f"Error loading settings: {e}")
                return [], [], {}
        return [], [], {}

    def update_excel_model_compatibility(self):
        """Update Excel file to match current type-to-model compatibility settings"""
        try:
            from openpyxl import load_workbook
            self.ensure_excel_file()  # Lazy load Excel file
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Track changes
            changes_made = 0
            
            # Iterate through all rows (skip header)
            for row_idx in range(2, ws.max_row + 1):
                serial = ws.cell(row=row_idx, column=1).value
                fixture_type = ws.cell(row=row_idx, column=3).value  # Name column (fixture type) - CORRECTED to column 3
                
                if fixture_type and serial:
                    # Get compatible models for this fixture type
                    compatible_models = self.type_to_models.get(fixture_type, [])
                    
                    # Update Model column (column 2) with compatible models - CORRECTED to column 2
                    new_models_value = ", ".join(compatible_models) if compatible_models else ""
                    current_value = ws.cell(row=row_idx, column=2).value or ""
                    
                    if str(current_value) != new_models_value:
                        ws.cell(row=row_idx, column=2).value = new_models_value
                        changes_made += 1
            
            # Save if changes were made
            if changes_made > 0:
                wb.save(self.excel_file)
                wb.close()
                print(f"Updated {changes_made} fixture(s) in Excel to match current model compatibility.")
            else:
                wb.close()
                print("No Excel updates needed - all fixtures already match current compatibility.")
                
        except Exception as e:
            print(f"Error updating Excel model compatibility: {str(e)}")

    def update_excel_names(self, model_name_changes, type_name_changes):
        """Update Excel file when model or type names are changed in settings"""
        try:
            from openpyxl import load_workbook
            self.ensure_excel_file()  # Lazy load Excel file
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            changes_made = 0
            
            # Iterate through all rows (skip header)
            for row_idx in range(2, ws.max_row + 1):
                # Column 1: Serial
                # Column 2: Model (compatible models, comma-separated)
                # Column 3: Name (fixture type)
                # Column 4: Status
                
                # Update Model column (column 2) - replace old model names with new ones
                if model_name_changes:
                    model_cell = ws.cell(row=row_idx, column=2)
                    if model_cell.value:
                        models_str = str(model_cell.value)
                        updated = False
                        # Split by comma, update each model name
                        models_list = [m.strip() for m in models_str.split(",")]
                        for old_name, new_name in model_name_changes.items():
                            if old_name in models_list:
                                models_list = [new_name if m == old_name else m for m in models_list]
                                updated = True
                        if updated:
                            model_cell.value = ", ".join(models_list)
                            changes_made += 1
                
                # Update Name column (column 3) - replace old type names with new ones
                if type_name_changes:
                    name_cell = ws.cell(row=row_idx, column=3)
                    if name_cell.value:
                        old_type = str(name_cell.value)
                        if old_type in type_name_changes:
                            name_cell.value = type_name_changes[old_type]
                            changes_made += 1
            
            # Save if changes were made
            if changes_made > 0:
                wb.save(self.excel_file)
                wb.close()
                print(f"Updated {changes_made} cell(s) in Excel with new names.")
            else:
                wb.close()
                print("No name changes needed in Excel.")
                
        except Exception as e:
            QMessageBox.warning(self, "Excel Update Error", 
                              f"Failed to update model compatibility in Excel:\n{str(e)}")
    
    def save_images_to_excel(self, type_images):
        """Save fixture type images to Excel file on a visible separate sheet with actual images"""
        try:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            from openpyxl.styles import Font, PatternFill, Alignment
            import base64
            from io import BytesIO
            from PIL import Image
            
            self.ensure_excel_file()
            wb = load_workbook(self.excel_file)
            
            # Create or recreate "Fixture Images" sheet (visible, not hidden)
            # Also check for old name "FixtureImages" (without space) for backward compatibility
            if "Fixture Images" in wb.sheetnames:
                wb.remove(wb["Fixture Images"])
            if "FixtureImages" in wb.sheetnames:
                wb.remove(wb["FixtureImages"])
            
            img_sheet = wb.create_sheet("Fixture Images")
            # Keep it visible so user can see images in Excel
            
            # Add headers with styling
            img_sheet['A1'] = 'Fixture Type'
            img_sheet['B1'] = 'Image'
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            img_sheet['A1'].fill = header_fill
            img_sheet['A1'].font = header_font
            img_sheet['B1'].fill = header_fill
            img_sheet['B1'].font = header_font
            
            # Set column widths
            img_sheet.column_dimensions['A'].width = 30
            img_sheet.column_dimensions['B'].width = 50
            
            row = 2
            for fixture_type, image_base64 in type_images.items():
                # Add fixture type name
                img_sheet[f'A{row}'] = fixture_type
                img_sheet[f'A{row}'].alignment = Alignment(vertical='center')
                
                # Add image
                try:
                    image_bytes = base64.b64decode(image_base64)
                    img = Image.open(BytesIO(image_bytes))
                    
                    # Resize image for Excel (smaller for better performance)
                    img.thumbnail((200, 200), Image.Resampling.LANCZOS)
                    
                    # Save to BytesIO
                    img_buffer = BytesIO()
                    img.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    # Create openpyxl image
                    excel_img = OpenpyxlImage(img_buffer)
                    excel_img.width = 200
                    excel_img.height = 200
                    
                    # Place image in cell B
                    img_sheet.add_image(excel_img, f'B{row}')
                    
                    # Set row height to fit image (in points, 1 point ≈ 1.33 pixels)
                    img_sheet.row_dimensions[row].height = 150
                    
                except Exception as e:
                    print(f"Error adding image for {fixture_type}: {e}")
                    img_sheet[f'B{row}'] = "[Image failed to load]"
                
                row += 1
            
            # If no images, add a helpful message
            if len(type_images) == 0:
                img_sheet['A2'] = 'No images uploaded yet'
                img_sheet['B2'] = 'Use Settings > Edit Models/Names > Fixture Types > Upload Image to add images'
                img_sheet['A2'].font = Font(italic=True, color="808080")
                img_sheet['B2'].font = Font(italic=True, color="808080")
                img_sheet.merge_cells('B2:B5')
            
            wb.save(self.excel_file)
            wb.close()
            
            if len(type_images) > 0:
                print(f"✅ Saved {len(type_images)} image(s) to 'Fixture Images' sheet in Excel")
            else:
                print(f"✅ Created 'Fixture Images' sheet (no images yet)")
            
        except Exception as e:
            print(f"❌ Error saving images to Excel: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def load_images_from_excel(self):
        """Load fixture type images from Excel file - Note: Images are already loaded from JSON"""
        # Images are now stored in JSON for app use, and in Excel sheet for viewing only
        # No need to load from Excel since JSON is the source of truth
        # The "Fixture Images" sheet in Excel is just for visual reference
        pass
        
    def create_inventory_tab(self):
        """Create the inventory/search tab"""
        inventory_widget = QWidget()
        layout = QVBoxLayout(inventory_widget)
        
        # Search section
        search_layout = QHBoxLayout()
        search_label = QLabel("Search:")
        search_label.setFont(self.modern_font)
        self.search_input = QLineEdit()
        self.search_input.setFont(self.modern_font)
        self.search_input.setPlaceholderText("Enter serial number, model, or name...")
        self.search_input.textChanged.connect(self.filter_inventory)

        status_label = QLabel("Status:")
        status_label.setFont(self.modern_font)
        self.status_filter = QComboBox()
        self.status_filter.setFont(self.modern_font)
        self.status_filter.addItems(["All Status", "Available", "Checked Out", "WIP", "Discontinued"])
        self.status_filter.currentTextChanged.connect(self.filter_inventory)

        refresh_btn = QPushButton("Refresh")
        refresh_btn.setFont(self.modern_font)
        refresh_btn.clicked.connect(self.load_data)

        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input, stretch=3)
        search_layout.addWidget(status_label)
        search_layout.addWidget(self.status_filter, stretch=1)
        search_layout.addWidget(refresh_btn)
            
        layout.addLayout(search_layout)
        
        # Inventory table
        self.inventory_table = QTableWidget()
        self.inventory_table.setColumnCount(5)
        self.inventory_table.setHorizontalHeaderLabels([
            "Serial", "Model", "Name", "Status", "WIP Location"
        ])
        self.inventory_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.inventory_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        # Allow manual column resizing
        self.inventory_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        # Set reasonable default column widths
        self.inventory_table.setColumnWidth(0, 150)  # Serial
        self.inventory_table.setColumnWidth(1, 400)  # Model (wider for multiple models)
        self.inventory_table.setColumnWidth(2, 300)  # Name
        self.inventory_table.setColumnWidth(3, 120)  # Status
        self.inventory_table.setAlternatingRowColors(True)
        self.inventory_table.setFont(self.modern_font)
        self.inventory_table.horizontalHeader().setFont(self.header_font)
        self.inventory_table.setWordWrap(True)
        # Enable text to wrap and rows to auto-adjust height
        self.inventory_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        
        layout.addWidget(self.inventory_table)
        
        # Statistics section
        stats_layout = QHBoxLayout()
        self.total_label = QLabel("Total Items: 0")
        self.total_label.setFont(self.header_font)
        self.available_label = QLabel("Available: 0")
        self.available_label.setFont(self.header_font)
        self.checked_out_label = QLabel("Checked Out: 0")
        self.checked_out_label.setFont(self.header_font)

        stats_layout.addWidget(self.total_label)
        stats_layout.addWidget(self.available_label)
        stats_layout.addWidget(self.checked_out_label)
        stats_layout.addStretch()
            
        layout.addLayout(stats_layout)
            
        # Delete button
        delete_btn = QPushButton("Delete Serial")
        delete_btn.setFont(self.modern_font)
        delete_btn.clicked.connect(self.delete_selected_serial)
        layout.addWidget(delete_btn)
        
        self.tabs.addTab(inventory_widget, "Inventory Search")
    
    def create_checkin_checkout_tab(self):
        """Create the check-in/check-out tab"""
        checkin_checkout_widget = QWidget()
        layout = QVBoxLayout(checkin_checkout_widget)
        
        # Serial number input
        serial_layout = QHBoxLayout()
        serial_label = QLabel("Serial Number:")
        serial_label.setFont(self.modern_font)
        self.serial_input = QLineEdit()
        self.serial_input.setFont(self.modern_font)
        self.serial_input.setPlaceholderText("Enter serial number...")
        lookup_btn = QPushButton("Look Up")
        lookup_btn.setFont(self.modern_font)
        lookup_btn.clicked.connect(self.lookup_fixture)

        serial_layout.addWidget(serial_label)
        serial_layout.addWidget(self.serial_input, stretch=2)
        serial_layout.addWidget(lookup_btn)
        
        layout.addLayout(serial_layout)
        
        # Fixture details display
        details_label = QLabel("Fixture Details:")
        details_label.setFont(self.modern_font)
        self.details_text = QTextEdit()
        self.details_text.setFont(self.modern_font)
        self.details_text.setReadOnly(True)
        self.details_text.setMaximumHeight(150)
        layout.addWidget(details_label)
        layout.addWidget(self.details_text)
        
        # Check-out section
        checkout_group = QWidget()
        checkout_layout = QFormLayout(checkout_group)
        
        self.checkout_person_input = QLineEdit()
        self.checkout_person_input.setFont(self.modern_font)
        self.checkout_person_input.setPlaceholderText("Enter name...")
        # Autofill with Windows username every time tab is opened
        self.tabs.currentChanged.connect(self.autofill_employee_name)

        checked_out_by_label = QLabel("Processed By:")
        checked_out_by_label.setFont(self.modern_font)
        checkout_layout.addRow(checked_out_by_label, self.checkout_person_input)
        
        # WIP Section
        wip_layout = QHBoxLayout()
        self.wip_checkbox = QCheckBox("Mark as WIP (Work In Progress)")
        self.wip_checkbox.setFont(self.modern_font)
        self.wip_checkbox.stateChanged.connect(self.toggle_wip_location)
        wip_layout.addWidget(self.wip_checkbox)
        
        self.wip_location_label = QLabel("WIP Location:")
        self.wip_location_label.setFont(self.modern_font)
        self.wip_location_label.setVisible(False)
        
        self.wip_location_combo = QComboBox()
        self.wip_location_combo.setFont(self.modern_font)
        self.wip_location_combo.setEditable(True)  # Searchable
        self.wip_location_combo.setVisible(False)
        # Load WIP locations
        wip_locations = getattr(self, 'wip_locations', ["Bonepile", "Debug", "Operations", "Eng Lab"])
        self.wip_location_combo.addItems(wip_locations)
        
        checkout_layout.addRow(wip_layout)
        checkout_layout.addRow(self.wip_location_label, self.wip_location_combo)
        
        check_out_label = QLabel("Fill Data:")
        check_out_label.setFont(self.modern_font)
        layout.addWidget(check_out_label)
        layout.addWidget(checkout_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        checkout_btn = QPushButton("Check Out Fixture")
        checkout_btn.setFont(self.header_font)
        checkout_btn.clicked.connect(self.checkout_fixture)
        checkout_btn.setStyleSheet("QPushButton { background-color: #d9534f; color: white; font-weight: bold; padding: 16px; border-radius: 8px; }")

        checkin_btn = QPushButton("Check In Fixture")
        checkin_btn.setFont(self.header_font)
        checkin_btn.clicked.connect(self.checkin_fixture)
        checkin_btn.setStyleSheet("QPushButton { background-color: #5cb85c; color: white; font-weight: bold; padding: 16px; border-radius: 8px; }")

        discontinued_btn = QPushButton("Mark as Discontinued")
        discontinued_btn.setFont(self.header_font)
        discontinued_btn.clicked.connect(self.discontinue_fixture)
        discontinued_btn.setStyleSheet("QPushButton { background-color: #6c757d; color: white; font-weight: bold; padding: 16px; border-radius: 8px; }")

        button_layout.addWidget(checkout_btn)
        button_layout.addWidget(checkin_btn)
        button_layout.addWidget(discontinued_btn)
            
        layout.addLayout(button_layout)
        layout.addStretch()
        
        self.tabs.addTab(checkin_checkout_widget, "Check In/Out")
    
    def create_serialize_tab(self):
        """Create the serialize new items tab"""
        serialize_widget = QWidget()
        layout = QVBoxLayout(serialize_widget)

        form_layout = QFormLayout()
        # Removed Name and Model dropdowns for serializer creation

        self.new_serial_input = QLineEdit()
        self.new_serial_input.setFont(self.modern_font)
        self.new_serial_input.setReadOnly(False)  # Allow editing
        self.new_serial_input.setPlaceholderText("Enter serial number (e.g., FX00001)")
        # Do NOT auto-fill serial - let user type manually

        self.type_dropdown = QComboBox()
        self.type_dropdown.setFont(self.modern_font)
        self.type_dropdown.setEditable(True)  # Allow typing to search
        self.type_dropdown.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)  # Don't add typed text as new item
        self.compatible_models_label = QLabel()
        self.compatible_models_label.setFont(self.modern_font)
        self.compatible_models_label.setWordWrap(True)

        def update_compatible_models_label():
            selected_type = self.type_dropdown.currentText()
            compatible_models = self.type_to_models.get(selected_type, []) if hasattr(self, 'type_to_models') else []
            if compatible_models:
                self.compatible_models_label.setText(", ".join(compatible_models))
            else:
                self.compatible_models_label.setText("(No compatible models)")
        self.type_dropdown.currentTextChanged.connect(update_compatible_models_label)

        def update_serializer_dropdowns():
            self.type_dropdown.clear()
            self.type_dropdown.addItem("")  # Add empty item at the start
            types = getattr(self, 'types', []) if hasattr(self, 'types') else []
            self.type_dropdown.addItems(types)
            self.type_dropdown.setCurrentIndex(0)  # Start with empty selection
            update_compatible_models_label()
        self.update_serializer_dropdowns = update_serializer_dropdowns
        update_serializer_dropdowns()

        serial_label = QLabel("Serial Number:*")
        serial_label.setFont(self.modern_font)
        type_label = QLabel("Fixture Type:*")
        type_label.setFont(self.modern_font)
        model_label = QLabel("Model (compatible):*")
        model_label.setFont(self.modern_font)

        form_layout.addRow(serial_label, self.new_serial_input)
        form_layout.addRow(type_label, self.type_dropdown)
        form_layout.addRow(model_label, self.compatible_models_label)

        layout.addLayout(form_layout)

        # Add button
        add_btn = QPushButton("Add New Fixture")
        add_btn.setFont(self.header_font)
        def add_and_reset_form():
            self.add_new_fixture()
            self.type_dropdown.setCurrentIndex(0)  # Reset to empty
            self.type_dropdown.clearEditText()     # Clear any typed text
            self.new_serial_input.clear()          # Clear serial input
            self.new_serial_input.setFocus()       # Auto-focus serial input for next entry
        add_btn.clicked.connect(add_and_reset_form)
        add_btn.setStyleSheet("QPushButton { background-color: #0275d8; color: white; font-weight: bold; padding: 16px; border-radius: 8px; }")

        layout.addWidget(add_btn)
        layout.addStretch()

        self.tabs.addTab(serialize_widget, "Serialize New Item")
        # Store reference for tab change handler to auto-focus serial input
        self.serialize_tab_widget = serialize_widget

    def refresh_excel_file(self):
        """Return the newest Excel file path available (for refreshing data)"""
        newest = self.get_newest_excel_file()
        if newest:
            print(f"Refresh: newest Excel is {os.path.basename(newest)}")
            return newest
        return self.excel_file

    def checkout_fixture(self):
        try:
            # Manage Excel files before saving
            if not self.manage_excel_files_on_save():
                QMessageBox.warning(self, "File Error", "Could not prepare Excel file for saving.")
                return
                
            serial = self.serial_input.text().strip().upper()
            person = self.checkout_person_input.text().strip()
            if not serial or not person:
                QMessageBox.warning(self, "Missing Info", "Serial number and employee name are required.")
                return
            # Validate serial syntax before proceeding
            import re
            if not re.match(r'^(FX\d{5}|F\d{2,6})(-SLC)?$', serial):
                QMessageBox.warning(self, "Input Error", "Serial must be in format FX##### or F## to F######, optionally with -SLC suffix. Checkout rejected.")
                return
            
            # Check WIP status
            is_wip = self.wip_checkbox.isChecked()
            wip_location = self.wip_location_combo.currentText().strip() if is_wip else ""
            
            if is_wip and not wip_location:
                QMessageBox.warning(self, "Missing Info", "Please select a WIP location.")
                return
            
            checkout_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            from openpyxl import load_workbook
            from openpyxl.styles import numbers
            wb = load_workbook(self.excel_file)
            ws = wb.active
            found = False
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if str(row[0].value) == serial:
                    # Check if fixture is available for checkout
                    current_status = str(row[3].value).strip() if row[3].value else ""
                    if current_status != "Available":
                        wb.close()
                        QMessageBox.warning(self, "Cannot Check Out", 
                                          f"Fixture '{serial}' cannot be checked out.\nCurrent status: {current_status}\n\nOnly fixtures with 'Available' status can be checked out.")
                        return
                    
                    # Set status based on WIP checkbox
                    if is_wip:
                        row[3].value = "WIP"
                    else:
                        row[3].value = "Checked Out"
                    
                    # Ensure columns exist
                    if ws.max_column < 8:
                        ws.cell(row=1, column=5, value="Checked Out By")
                        ws.cell(row=1, column=6, value="Checked Out At")
                        ws.cell(row=1, column=8, value="WIP Location")
                    
                    ws.cell(row=row_idx, column=5, value=person)
                    # Store as text string, not as date
                    checkout_cell = ws.cell(row=row_idx, column=6)
                    checkout_cell.value = checkout_time
                    checkout_cell.number_format = '@'  # Text format
                    
                    # Save WIP location if applicable
                    if is_wip:
                        ws.cell(row=row_idx, column=8, value=wip_location)
                    else:
                        ws.cell(row=row_idx, column=8, value="")
                    
                    found = True
                    break
            wb.save(self.excel_file)
            wb.close()
            print(f"Fixture {serial} checked out in {self.excel_file}")
            
            # Update details display
            status_text = "WIP" if is_wip else "Checked Out"
            details = f"Serial: {serial}\nStatus: {status_text}\nChecked Out By: {person}\nChecked Out At: {checkout_time}"
            if is_wip:
                details += f"\nWIP Location: {wip_location}"
            self.details_text.setText(details)
            
            # Show warning if not found (workbook already closed above)
            if not found:
                QMessageBox.warning(self, "Not Found", f"Serial '{serial}' not found in inventory.")
                return
            
            self.load_data()
            self.serial_input.clear()
            self.checkout_person_input.clear()
            self.wip_checkbox.setChecked(False)  # Reset WIP checkbox
        except Exception as e:
            QMessageBox.critical(self, "Checkout Error", f"An error occurred during checkout:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def checkin_fixture(self):
        # Manage Excel files before saving
        if not self.manage_excel_files_on_save():
            QMessageBox.warning(self, "File Error", "Could not prepare Excel file for saving.")
            return
            
        serial = self.serial_input.text().strip().upper()
        person = self.checkout_person_input.text().strip()
        if not serial or not person:
            QMessageBox.warning(self, "Missing Info", "Serial number and employee name are required.")
            return
        # Validate serial syntax before proceeding
        import re
        if not re.match(r'^(FX\d{5}|F\d{2,6})(-SLC)?$', serial):
            QMessageBox.warning(self, "Input Error", "Serial must be in format FX##### or F## to F######, optionally with -SLC suffix. Check-in rejected.")
            return
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_file)
            ws = wb.active
            found = False
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == serial:
                    current_status = ws.cell(row=r, column=4).value
                    if current_status not in ["Checked Out", "WIP"]:
                        QMessageBox.warning(self, "Not Checked Out", f"Serial {serial} is not checked out or in WIP status.")
                        wb.close()
                        return
                    ws.cell(row=r, column=4, value="Available")
                    # Clear WIP location
                    ws.cell(row=r, column=8, value="")
                    wb.save(self.excel_file)
                    wb.close()
                    print(f"Fixture {serial} checked in to {self.excel_file}")
                    found = True
                    break
            
            # Close workbook before showing message boxes
            if not found:
                wb.close()
                QMessageBox.warning(self, "Not Found", f"No fixture found with serial number: {serial}")
                return
            QMessageBox.information(self, "Success", f"Fixture {serial} checked in.")
            self.serial_input.clear()
            self.checkout_person_input.clear()
            self.details_text.clear()
            self.load_data()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to check in fixture: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def discontinue_fixture(self):
        """Mark a fixture as discontinued"""
        # Manage Excel files before saving
        if not self.manage_excel_files_on_save():
            QMessageBox.warning(self, "File Error", "Could not prepare Excel file for saving.")
            return
            
        serial = self.serial_input.text().strip().upper()
        person = self.checkout_person_input.text().strip()
        if not serial or not person:
            QMessageBox.warning(self, "Missing Info", "Serial number and employee name are required.")
            return
        # Validate serial syntax before proceeding
        import re
        if not re.match(r'^(FX\d{5}|F\d{2,6})(-SLC)?$', serial):
            QMessageBox.warning(self, "Input Error", "Serial must be in format FX##### or F## to F######, optionally with -SLC suffix.")
            return
        
        # Confirm discontinuation
        from PyQt6.QtWidgets import QMessageBox
        reply = QMessageBox.question(self, "Confirm Discontinuation", 
                                     f"Are you sure you want to mark fixture '{serial}' as DISCONTINUED?\n\nThis action indicates the fixture is no longer in service.",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_file)
            ws = wb.active
            found = False
            discontinued_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == serial:
                    # Set status to Discontinued
                    ws.cell(row=r, column=4, value="Discontinued")
                    # Record who discontinued it and when
                    ws.cell(row=r, column=5, value=person)
                    ws.cell(row=r, column=6).value = discontinued_time
                    ws.cell(row=r, column=6).number_format = '@'  # Text format
                    # Clear WIP location if any
                    ws.cell(row=r, column=8, value="")
                    
                    wb.save(self.excel_file)
                    wb.close()
                    print(f"Fixture {serial} marked as discontinued in {self.excel_file}")
                    found = True
                    break
            
            # Close workbook before showing message boxes
            if not found:
                wb.close()
                QMessageBox.warning(self, "Not Found", f"No fixture found with serial number: {serial}")
                return
            
            QMessageBox.information(self, "Success", f"Fixture {serial} has been marked as Discontinued.")
            self.details_text.setText(f"Serial: {serial}\nStatus: Discontinued\nProcessed By: {person}\nDate: {discontinued_time}")
            self.serial_input.clear()
            self.checkout_person_input.clear()
            self.load_data()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to discontinue fixture: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def add_new_fixture(self):
        """Add a new fixture to inventory"""
        # Manage Excel files before saving
        if not self.manage_excel_files_on_save():
            QMessageBox.warning(self, "File Error", "Could not prepare Excel file for saving.")
            return
            
        serial = self.new_serial_input.text().strip().upper()
        fixture_type = self.type_dropdown.currentText().strip()
        compatible_models = self.compatible_models_label.text().strip()
        status = 'Available'
        serialized_date = datetime.now().strftime('%Y-%m-%d')

        import re
        if not serial or not serial.strip() or not fixture_type:
            QMessageBox.warning(self, "Input Error", "Serial number and fixture type are required.")
            return
        
        # Validate that fixture type is registered
        registered_types = getattr(self, 'types', [])
        if fixture_type not in registered_types:
            QMessageBox.warning(self, "Invalid Fixture Type", 
                              f"'{fixture_type}' is not a registered fixture type.\n\n"
                              f"Please select a fixture type from the dropdown or add it in Settings → Edit Models/Names.")
            return
        
        # Accept formats: FX##### (5 digits), F## to F###### (2-6 digits), optionally with -SLC
        if not re.match(r'^(FX\d{5}|F\d{2,6})(-SLC)?$', serial):
            QMessageBox.warning(self, "Input Error", "Serial must be in format FX##### or F## to F######, optionally with -SLC suffix.")
            return

        try:
            from openpyxl import load_workbook
            print(f"add_new_fixture: Working with file: {self.excel_file}")
            wb = load_workbook(self.excel_file)
            ws = wb.active
            # Check for duplicate serial
            serials = [str(cell.value) for cell in ws['A'] if cell.row != 1]
            if serial in serials:
                QMessageBox.warning(self, "Duplicate Serial", f"A fixture with serial number {serial} already exists.")
                return
            # Find next empty row
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=serial)
            ws.cell(row=next_row, column=2, value=compatible_models)  # Model column
            ws.cell(row=next_row, column=3, value=fixture_type)       # Name column
            ws.cell(row=next_row, column=4, value=status)
            ws.cell(row=next_row, column=5, value=None)  # Checked Out By
            ws.cell(row=next_row, column=6, value=None)  # Checked Out At
            ws.cell(row=next_row, column=7, value=serialized_date)
            wb.save(self.excel_file)
            wb.close()
            print(f"Fixture {serial} saved to {self.excel_file}")
            QMessageBox.information(self, "Success", f"Fixture {serial} added successfully.")
            # Don't auto-fill next serial - let user type manually
            self.load_data()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to add fixture: {str(e)}")
            import traceback
            traceback.print_exc()
            # Ensure workbook is closed even on error
            try:
                wb.close()
            except:
                pass

    def delete_selected_serial(self):
        # Manage Excel files before saving
        if not self.manage_excel_files_on_save():
            QMessageBox.warning(self, "File Error", "Could not prepare Excel file for saving.")
            return
            
        selected_rows = self.inventory_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Delete Serial", "Please select a serial to delete.")
            return
        serials_to_delete = [self.inventory_table.item(row.row(), 0).text() for row in selected_rows]
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_file)
            ws = wb.active
            rows_to_keep = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[0]) not in serials_to_delete:
                    rows_to_keep.append(row)
            ws.delete_rows(2, ws.max_row)
            wb.save(self.excel_file)
            for row_data in rows_to_keep:
                ws.append(row_data)
            wb.save(self.excel_file)
            wb.close()
            print(f"Deleted {len(serials_to_delete)} serial(s) from {self.excel_file}")
            self.load_data()
            QMessageBox.information(self, "Delete Serial", f"Deleted serial(s): {', '.join(serials_to_delete)}")
        except Exception as e:
            QMessageBox.critical(self, "Delete Error", f"An error occurred during delete:\n{str(e)}")
            import traceback
            traceback.print_exc()
            # Ensure workbook is closed even on error
            try:
                wb.close()
            except:
                pass
    
    def lookup_fixture(self):
        query = self.serial_input.text().strip()
        if not query:
            QMessageBox.warning(self, "Missing Info", "Please enter a serial number or fixture name to look up.")
            return
        query_upper = query.upper()
        query_lower = query.lower()
        try:
            from openpyxl import load_workbook
            from datetime import datetime
            # Always use newest Excel for lookup
            newest_excel = self.get_newest_excel_file()
            if newest_excel is None:
                QMessageBox.warning(self, "No Data", "No Excel inventory files found.")
                return
            print(f"Lookup using newest Excel: {os.path.basename(newest_excel)}")
            wb = load_workbook(newest_excel, read_only=True)
            ws = wb.active
            matches = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                serial_cell = '' if row[0] is None else str(row[0])
                name_cell = '' if len(row) > 2 and row[2] is not None else ''
                if len(row) > 2 and row[2] is not None:
                    name_cell = str(row[2])

                serial_match = (serial_cell.upper() == query_upper)
                name_match = (query_lower in name_cell.lower()) if name_cell else False

                if serial_match or name_match:
                    details = []
                    details.append(f"Serial: {serial_cell}")
                    details.append(f"Model: {row[1] if len(row) > 1 and row[1] is not None else ''}")
                    details.append(f"Name: {name_cell}")
                    status = row[3] if len(row) > 3 and row[3] is not None else ''
                    details.append(f"Status: {status}")
                    
                    # Show WIP location if status is WIP
                    if status == "WIP" and len(row) > 7 and row[7]:
                        details.append(f"WIP Location: {row[7]}")
                    
                    # Type description (from settings)
                    desc = getattr(self, 'type_descriptions', {}).get(name_cell, '')
                    if desc:
                        details.append(f"Type Description: {desc}")
                    # Optional: show who checked out and when
                    if len(row) > 4 and row[4]:
                        details.append(f"Checked Out By: {row[4]}")
                    if len(row) > 5 and row[5]:
                        checkout_time = row[5]
                        if isinstance(checkout_time, datetime):
                            checkout_time = checkout_time.strftime('%Y-%m-%d %H:%M:%S')
                        details.append(f"Checked Out At: {checkout_time}")
                    matches.append("\n".join(details))

            wb.close()
            if matches:
                # If multiple fixtures match, separate them for clarity
                out = "\n\n---\n\n".join(matches)
                self.details_text.setText(out)
            else:
                QMessageBox.information(self, "Not Found", f"No fixtures matching '{query}' were found in inventory.")
        except Exception as e:
            QMessageBox.critical(self, "Lookup Error", f"An error occurred during lookup:\n{str(e)}")
    
    def autofill_employee_name(self, index):
        # Check if the checkin/checkout tab is selected
        if self.tabs.tabText(index) == "Check In/Out":
            self.checkout_person_input.setText(self.username)

    def toggle_wip_location(self, state):
        """Show/hide WIP location dropdown based on checkbox state"""
        is_checked = (state == 2)  # Qt.CheckState.Checked == 2
        self.wip_location_label.setVisible(is_checked)
        self.wip_location_combo.setVisible(is_checked)

    def eventFilter(self, source, event):
        from PyQt6.QtCore import QEvent
        if hasattr(self, 'inventory_table') and source == self.inventory_table.viewport():
            if event.type() == QEvent.Type.MouseButtonPress:
                # Don't clear selection, just allow normal clicking behavior
                pass
        return super().eventFilter(source, event)

    def filter_inventory(self):
        search_text = self.search_input.text().lower()
        status_filter = self.status_filter.currentText()
        terms = [t.strip() for t in search_text.split(',') if t.strip()]
        filtered_count = 0
        available_count = 0
        checked_out_count = 0
        wip_count = 0
        discontinued_count = 0
        for row in range(self.inventory_table.rowCount()):
            show_row = True
            if terms:
                row_match = True
                for term in terms:
                    found = False
                    for col in [0, 1, 2]:
                        item = self.inventory_table.item(row, col)
                        if item and term in item.text().lower():
                            found = True
                            break
                    row_match = row_match and found
                show_row = show_row and row_match
            if status_filter != "All Status":
                status_item = self.inventory_table.item(row, 3)
                if status_item:
                    show_row = show_row and (status_item.text() == status_filter)
            self.inventory_table.setRowHidden(row, not show_row)
            if show_row:
                filtered_count += 1
                status_item = self.inventory_table.item(row, 3)
                if status_item:
                    if status_item.text() == "Available":
                        available_count += 1
                    elif status_item.text() == "Checked Out":
                        checked_out_count += 1
                    elif status_item.text() == "WIP":
                        wip_count += 1
                    elif status_item.text() == "Discontinued":
                        discontinued_count += 1
        self.total_label.setText(f"Total Items: {filtered_count}")
        self.available_label.setText(f"Available: {available_count}")
        self.checked_out_label.setText(f"Checked Out: {checked_out_count} | WIP: {wip_count} | Discontinued: {discontinued_count}")

    def auto_refresh(self):
        """Auto-refresh both Excel data and settings - only if files have changed"""
        try:
            needs_reload = False
            
            # Check if settings file has changed
            newest_settings = self.get_newest_settings_file()
            if newest_settings and os.path.exists(newest_settings):
                settings_mtime = os.path.getmtime(newest_settings)
                if settings_mtime > self._last_settings_mtime:
                    self._last_settings_mtime = settings_mtime
                    # Store old settings to detect changes
                    old_types = getattr(self, 'types', [])
                    old_models = getattr(self, 'models', [])
                    
                    self.settings_file = newest_settings
                    self.models, self.types, self.type_to_models = self.load_settings()
                    
                    # Only update serializer dropdowns if settings actually changed
                    if (old_types != self.types or old_models != self.models) and hasattr(self, 'update_serializer_dropdowns'):
                        # Preserve the current serial input value
                        current_serial = self.new_serial_input.text() if hasattr(self, 'new_serial_input') else ""
                        self.update_serializer_dropdowns()
                        # Restore serial input if user was typing
                        if current_serial and hasattr(self, 'new_serial_input'):
                            self.new_serial_input.setText(current_serial)
                    needs_reload = True
            
            # Check if Excel file has changed
            newest_excel = self.get_newest_excel_file()
            if newest_excel and os.path.exists(newest_excel):
                excel_mtime = os.path.getmtime(newest_excel)
                if excel_mtime > self._last_excel_mtime:
                    self._last_excel_mtime = excel_mtime
                    needs_reload = True
            
            # Only reload inventory if files actually changed
            if needs_reload:
                self.load_data()
        except Exception as e:
            # Silently fail on auto-refresh to avoid popup spam
            pass

    def load_data(self):
        try:
            # Always load from the newest Excel file available
            newest_excel = self.get_newest_excel_file()
            if newest_excel is None:
                print(f"No Excel files found in {self.base_dir}")
                self.inventory_table.setRowCount(0)
                return
            
            print(f"Loading data from newest Excel: {os.path.basename(newest_excel)}")
            excel_to_read = newest_excel
            
            # Try to load 'Inventory' sheet, else load first sheet
            # Use openpyxl engine with read_only for better performance
            try:
                df = pd.read_excel(excel_to_read, sheet_name="Inventory", engine='openpyxl')
            except Exception:
                xl = pd.ExcelFile(excel_to_read, engine='openpyxl')
                df = xl.parse(xl.sheet_names[0])
                xl.close()  # Close ExcelFile object

            # Normalize column names (strip, lower, remove spaces, dashes, underscores, parentheses)
            import re
            original_cols = df.columns.tolist()
            norm = lambda c: re.sub(r'[\s\-_()]', '', str(c).strip().lower())
            normalized_cols = [norm(c) for c in df.columns]
            
            print(f"Original columns: {original_cols}")
            print(f"Normalized columns: {normalized_cols}")
            
            col_map = {}
            col_to_original = {}
            for i, col in enumerate(normalized_cols):
                if col == 'serial' and 'Serial' not in col_map:
                    col_map['Serial'] = i
                    col_to_original['Serial'] = original_cols[i]
                elif 'model' in col and 'Model' not in col_map:
                    col_map['Model'] = i
                    col_to_original['Model'] = original_cols[i]
                elif 'name' in col and 'Name' not in col_map:
                    col_map['Name'] = i
                    col_to_original['Name'] = original_cols[i]
                elif 'status' in col and 'Status' not in col_map:
                    col_map['Status'] = i
                    col_to_original['Status'] = original_cols[i]
                elif ('shelf' in col or 'serializeddate' in col or 'shelfdate' in col or (col.endswith('date') and col != 'serial')) and 'Shelf-date' not in col_map:
                    col_map['Shelf-date'] = i
                    col_to_original['Shelf-date'] = original_cols[i]
                elif ('wip' in col and 'location' in col) and 'WIP-Location' not in col_map:
                    col_map['WIP-Location'] = i
                    col_to_original['WIP-Location'] = original_cols[i]
            
            print(f"Column mapping: {col_to_original}")
            
            required_cols = ['Serial', 'Model', 'Name', 'Status']
            if not all(k in col_map for k in required_cols):
                print(f"Excel file missing required columns. Found: {list(col_to_original.keys())}")
                self.inventory_table.setRowCount(0)
                return
            
            # Display as-is from Excel: Serial, Model, Name, Status, and Shelf-date/WIP Location if available
            display_cols = [col_to_original[k] for k in required_cols]
            if 'Shelf-date' in col_to_original:
                display_cols.append(col_to_original['Shelf-date'])
            
            # Get WIP Location column from original DataFrame and add to display
            wip_location_col_idx = col_map.get('WIP-Location', None)
            wip_location_series = None
            if wip_location_col_idx is not None:
                wip_location_col_name = original_cols[wip_location_col_idx]
                wip_location_series = df[wip_location_col_name]
            
            df_display = df[display_cols]
            self.inventory_table.setRowCount(0)
            for index, row in df_display.iterrows():
                row_position = self.inventory_table.rowCount()
                self.inventory_table.insertRow(row_position)
                
                # Get shelf-date value for this row if available
                shelf_date_val = None
                if 'Shelf-date' in col_to_original and len(display_cols) > 4:
                    shelf_date_val = row.iloc[4] if pd.notna(row.iloc[4]) else None
                
                # Get WIP Location value efficiently from pre-extracted series
                wip_location_val = ""
                if wip_location_series is not None:
                    wip_val = wip_location_series.iloc[index]
                    wip_location_val = str(wip_val) if pd.notna(wip_val) else ""
                
                for col_index, value in enumerate(row):
                    # For Model column (index 1 in display), list models vertically
                    if col_index == 1 and pd.notna(value):
                        models = [m.strip() for m in str(value).split(",")]
                        item = QTableWidgetItem("\n".join(models))
                    else:
                        item = QTableWidgetItem(str(value) if pd.notna(value) else "")
                    
                    # Apply cell coloring based on status and shelf life
                    if col_index == 3:  # Status column
                        status_val = str(value).strip()
                        
                        # Check shelf life if status is Available and we have a shelf date
                        if status_val == "Available" and shelf_date_val is not None:
                            try:
                                # Convert shelf_date_val to string for parsing
                                date_str = str(shelf_date_val)
                                serialized_dt = None
                                
                                # Try parsing as YYYY-MM-DD first
                                try:
                                    serialized_dt = datetime.strptime(date_str, "%Y-%m-%d")
                                except ValueError:
                                    # Try parsing as MM/DD/YYYY
                                    try:
                                        serialized_dt = datetime.strptime(date_str, "%m/%d/%Y")
                                    except ValueError:
                                        # Try pandas to_datetime for other formats
                                        serialized_dt = pd.to_datetime(date_str)
                                
                                if serialized_dt and (datetime.now() - serialized_dt).days >= 182:
                                    item.setText("Check")
                                    item.setBackground(QColor(255, 255, 0, 100))
                                    print(f"Fixture {row.iloc[0]} marked as Check (shelf life expired)")
                                else:
                                    item.setBackground(QColor(92, 184, 92, 50))
                            except Exception as e:
                                print(f"Error parsing date for {row.iloc[0]}: {e}")
                                item.setBackground(QColor(92, 184, 92, 50))
                        elif status_val == "Available":
                            item.setBackground(QColor(92, 184, 92, 50))
                        elif status_val == "Checked Out":
                            item.setBackground(QColor(217, 83, 79, 50))
                        elif status_val == "WIP":
                            # Blue color for WIP status
                            item.setBackground(QColor(52, 152, 219, 100))  # Nice blue color
                        elif status_val == "Discontinued":
                            # Gray color for Discontinued status
                            item.setBackground(QColor(108, 117, 125, 100))  # Dark gray color
                    
                    self.inventory_table.setItem(row_position, col_index, item)
                
                # Add WIP Location to column 4 (5th column)
                wip_item = QTableWidgetItem(wip_location_val)
                self.inventory_table.setItem(row_position, 4, wip_item)
            self.filter_inventory()
        except Exception as e:
            print(f"Error loading data: {e}")
            # Don't show error dialog on startup, just clear table
            self.inventory_table.setRowCount(0)

    def get_next_serial(self):
        from openpyxl import load_workbook
        self.ensure_excel_file()  # Get Excel file path
        
        # If file doesn't exist yet, return default starting serial
        if not os.path.exists(self.excel_file):
            return "FX00001"
        
        try:
            used = set()
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                val = str(row[0]) if row[0] is not None else ""
                if val.startswith('FX') and val[2:].isdigit():
                    used.add(int(val[2:]))
            wb.close()
            n = 1
            while n in used:
                n += 1
            return f"FX{n:05d}"
        except Exception as e:
            print(f"Error in get_next_serial: {e}")
            # Return default if any error occurs
            return "FX00001"

    def open_barcode_menu(self):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QLineEdit, QPushButton, QComboBox, QHBoxLayout
        import io
        import barcode
        from barcode.writer import ImageWriter
        from PIL import Image
        import tempfile
        import os
        dialog = QDialog(self)
        dialog.setWindowTitle("Generate and Print Barcode")
        dialog.resize(400, 250)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("Select or Search Serial Number for Barcode:"))
        # Load serials from Excel
        serials = []
        try:
            from openpyxl import load_workbook
            self.ensure_excel_file()  # Lazy load Excel file
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:
                    serials.append(str(row[0]))
            wb.close()
        except Exception:
            pass
        serial_combo = QComboBox()
        serial_combo.addItems(serials)
        layout.addWidget(serial_combo)
        search_input = QLineEdit()
        search_input.setPlaceholderText("Search serial...")
        def filter_serials():
            text = search_input.text().strip().lower()
            serial_combo.clear()
            filtered = [s for s in serials if text in s.lower()]
            serial_combo.addItems(filtered)
        search_input.textChanged.connect(filter_serials)
        layout.addWidget(search_input)
        barcode_label = QLabel()
        layout.addWidget(barcode_label)
        def generate_barcode():
            serial = serial_combo.currentText().strip()
            if not serial:
                barcode_label.setText("Please select a serial number.")
                return
            try:
                # Configure barcode writer for 60mm x 15mm label (at 200 DPI)
                # 60mm = ~472 pixels at 200 DPI, 15mm = ~118 pixels at 200 DPI
                # 30mm left offset = ~236 pixels at 200 DPI
                writer = ImageWriter()
                writer.set_options({
                    'module_width': 0.3,   # Width of bars in mm
                    'module_height': 12.0, # Height of bars in mm
                    'quiet_zone': 2.0,     # Margins in mm
                    'font_size': 10,       # Font size for text
                    'text_distance': 1.0,  # Distance between barcode and text
                    'dpi': 200             # Match printer DPI
                })
                ean = barcode.get('code128', serial, writer=writer)
                temp_dir = tempfile.gettempdir()
                os.makedirs(temp_dir, exist_ok=True)
                with tempfile.NamedTemporaryFile(suffix='.png', dir=temp_dir, delete=False) as tmp_file:
                    barcode_file = tmp_file.name
                # Save barcode
                ean.save(barcode_file[:-4])  # Remove .png as library adds it
                
                # Resize to exact 60mm x 15mm and add 30mm left margin using PIL
                from PIL import Image
                img = Image.open(barcode_file)
                # Calculate dimensions in pixels at 200 DPI
                target_width_px = int(60 * 200 / 25.4)   # ~472 pixels
                target_height_px = int(15 * 200 / 25.4)  # ~118 pixels
                left_margin_px = int(0 * 200 / 25.4)    # ~236 pixels
                
                # Resize barcode to fit label
                img_resized = img.resize((target_width_px, target_height_px), Image.Resampling.LANCZOS)
                
                # Create new image with left margin
                final_width = target_width_px + left_margin_px
                new_img = Image.new('RGB', (final_width, target_height_px), 'white')
                new_img.paste(img_resized, (left_margin_px, 0))
                new_img.save(barcode_file)
                
                barcode_label.setText(f"Barcode generated for {serial} (60mm x 15mm, 30mm left offset).")
                # Only print, do not open the PNG
                os.startfile(barcode_file, "print")
            except Exception as e:
                barcode_label.setText(f"Error: {str(e)}")
        gen_btn = QPushButton("Generate Barcode")
        gen_btn.clicked.connect(generate_barcode)
        layout.addWidget(gen_btn)
        dialog.exec()

def apply_dark_theme(app):
    """Apply dark theme to the application"""
    app.setStyle("Fusion")
    
    dark_palette = QPalette()
    dark_palette.setColor(QPalette.ColorRole.Window, QColor(53, 53, 53))
    dark_palette.setColor(QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
    dark_palette.setColor(QPalette.ColorRole.Base, QColor(35, 35, 35))
    dark_palette.setColor(QPalette.ColorRole.AlternateBase, QColor(53, 53, 53))
    dark_palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(25, 25, 25))
    dark_palette.setColor(QPalette.ColorRole.ToolTipText, Qt.GlobalColor.white)
    dark_palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
    dark_palette.setColor(QPalette.ColorRole.Button, QColor(53, 53, 53))
    dark_palette.setColor(QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
    dark_palette.setColor(QPalette.ColorRole.BrightText, Qt.GlobalColor.red)
    dark_palette.setColor(QPalette.ColorRole.Link, QColor(42, 130, 218))
    dark_palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
    dark_palette.setColor(QPalette.ColorRole.HighlightedText, Qt.GlobalColor.black)
    
    app.setPalette(dark_palette)


def main():
    app = QApplication(sys.argv)
    apply_dark_theme(app)
    
    window = FixtureControlApp()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
